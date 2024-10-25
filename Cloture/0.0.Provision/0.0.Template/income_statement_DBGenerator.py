from openpyxl import *
from random import randint, uniform
import datetime
import locale

# Définir la locale en français pour obtenir les mois en français
locale.setlocale(locale.LC_TIME, "fr_FR")

ENTREPRISE_NAME = "Entreprise fictive"
EMPLOYEES = 5000
NbEmployee = 0

TEMP_WORK = randint(round(EMPLOYEES * 0.07), round(EMPLOYEES * 0.12))
TEMP_WORK_5_LAST_YEARS = TEMP_WORK * 12 * 5
TEMP_WORK_COST = randint(12, 13) * 140 * uniform(1.5, 1.75)
NbTempWork = 0
nbOfLine = 1

ELECTRICITY_CONSUMPTION = 100000
GAZ_CONSUMPTION = 30000

ELECTRICITY_COST = [0.1673935, 0.1769825, 0.18188, 0.389, 0.2189865, 0.2516]
GAZ_COST = [0.0598, 0.5005, 0.05795, 0.0673, 0.1043, 0.1095]

dictionnaire = {
    "Personnel intérimaire": "Interim",
    "Medecine du travail, pharmacie": "Pole sante",
    "Fournitures non stockables (eau, énergie) - Eau": "Eau",
    "Fournitures non stockables (eau, énergie) - Gaz": "Gaz",
    "Fournitures non stockables (eau, énergie) - Electricité": "Electricite",
    "Sous-traitance générale - Cantine": "Cantine",
    "Entretien et réparations sur biens immobiliers": "Nettoyage"
}

CurrentDay = datetime.date.today()

def lastDayOfPreviousMounth(date):
    if str(date).split("-"):
        date = date.replace(day=1)
        last_month = date - datetime.timedelta(days=1)
        return last_month

def format_date(date):
    return date.strftime("%Y/%m/%d")

retroActiveDay = lastDayOfPreviousMounth(CurrentDay)

COST_CENTER = {
    "6200": "Usine",
    "6201": "Montage",
    "6202": "Usinage",
    "6203": "Magasin",
    "6204": "Expédition",
    "6205": "Reception",
    "6206": "Restauration",
    "6207": "Propreté",
    "6208": "Logistique"
}

PROFIT_CENTER = {
    "7200" : "Usine",
    "7201" : "Montage",
    "7202" : "Usinage",
    "7203" : "Magasin",
    "7204" : "Expédition",
    "7205" : "Reception",
    "7206" : "Restauration",
    "7207" : "Propreté",
    "7208" : "Logistique"
}

text_month = {
    "1" : "Janvier",
    "2" : "Février",
    "3" : "Mars",
    "4" : "Avril",
    "5" : "Mai",
    "6" : "Juin",
    "7" : "Juillet",
    "8" : "Août",
    "9" : "Septembre",
    "10" : "Octobre",
    "11" : "Novembre",
    "12" : "Décembre",
}

income_statement_DB_file = Workbook();
DB_ws = income_statement_DB_file.active;
DB_ws.title = "DB";

DB_ws["A1"] = "Nature comptable";
DB_ws.column_dimensions["A"].width = len("Nature comptable") + 1;
for i in range(2, (TEMP_WORK * 12 * 5) + 2): 
    DB_ws.cell(row = i, column = 1).value = "62110001";

DB_ws["B1"] = "Designation comptable";
DB_ws.column_dimensions["B"].width = len("Fournitures non stockables (eau, énergie) - Eau") + 1;
for i in range(2, (TEMP_WORK * 12 * 5) + 2): 
    DB_ws.cell(row = i, column = 2).value = "Personnel intérimaire";


DB_ws["C1"] = "Centre de coût";
DB_ws.column_dimensions["C"].width = len("Centre de coût") + 1;
for i in range(2, TEMP_WORK_5_LAST_YEARS + 2): 
    DB_ws.cell(row = i, column = 3).value = randint(6200, 6208);

DB_ws["D1"] = "Designation centre de coût";
DB_ws.column_dimensions["D"].width = len("Designation centre de coût") + 1;
# # Saisie des désignations des centres de coûts des 5 dernières années
# # En fonction des centres de coûts
a = 0
for i in range(2, (TEMP_WORK * 12 * 5) + 2): 
    for ii in range(1, len(COST_CENTER)) :

        cellValue = str(DB_ws.cell(row = i, column = 3).value);

        if(COST_CENTER[cellValue]) : 
            DB_ws.cell(row = i, column = 4).value = COST_CENTER[cellValue];
        


DB_ws["E1"] = "Centre de profit";
DB_ws.column_dimensions["E"].width = len("Centre de profit") + 1;
DB_ws["F1"] = "Designation centre de profit";
DB_ws.column_dimensions["F"].width = len("Designation centre de profit") + 1;

DB_ws["G1"] = "Montant";
DB_ws.column_dimensions["G"].width = len("Montant") + 1;
for i in range(2, TEMP_WORK_5_LAST_YEARS + 2): 
    DB_ws.cell(row = i, column = 7).value = TEMP_WORK_COST;

DB_ws["H1"] = "Type Piece";
DB_ws.column_dimensions["H"].width = len("Type Piece") + 1;
for i in range(2, TEMP_WORK_5_LAST_YEARS + 2): 
    DB_ws.cell(row = i, column = 8).value = "Facture";


DB_ws["I1"] = "Nom";
DB_ws.column_dimensions["I"].width = len("Nom") + 7;
DB_ws["J1"] = "Prenom";
DB_ws.column_dimensions["J"].width = len("Prenom") + 7;
DB_ws["K1"] = "Matricule";
DB_ws.column_dimensions["K"].width = len("Matricule") + 1;
DB_ws["L1"] = "Periode d'effet";
DB_ws.column_dimensions["L"].width = len("Periode d'effet") + 1;
DB_ws["M1"] = "Debut periode";
DB_ws.column_dimensions["M"].width = len("Debut periode") + 1;
DB_ws["N1"] = "Fin periode";
DB_ws.column_dimensions["N"].width = len("Fin periode") + 1;

for i in range(2, TEMP_WORK_5_LAST_YEARS + 2): 
    if NbTempWork < TEMP_WORK : 
        NbTempWork = NbTempWork + 1
    else:
        NbTempWork = 1
        retroActiveDay = lastDayOfPreviousMounth(retroActiveDay)

    DB_ws.cell(row = i, column = 9).value = "Nom" + str(NbTempWork);
    DB_ws.cell(row = i, column = 10).value = "Prenom" + str(NbTempWork);
    DB_ws.cell(row = i, column = 11).value = 100000 + int(NbTempWork);
    DB_ws.cell(row = i, column = 12).value = str(retroActiveDay).replace('-', '/');
    DB_ws.cell(row = i, column = 13).value = retroActiveDay.replace(day = 1);
    DB_ws.cell(row = i, column = 13).value = str(retroActiveDay).replace('-', '/');
    DB_ws.cell(row = i, column = 14).value = str(retroActiveDay).replace('-', '/');
    DB_ws.cell(row = i, column = 17).value = str(retroActiveDay).replace('-', '/');
    DB_ws.cell(row = i, column = 18).value = str(retroActiveDay).replace('-', '/');
    DB_ws.cell(row = i, column = 19).value = str(retroActiveDay).replace('-', '/');

DB_ws["O1"] = "N° piece reference";
DB_ws.column_dimensions["O"].width = len("N° piece reference") + 1;
for i in range(2, TEMP_WORK_5_LAST_YEARS + 2): 
    DB_ws.cell(row = i, column =15).value = "0000000000" + str(i-1);
    nbOfLine = nbOfLine + 1

DB_ws["P1"] = "Utilisateur ecriture";
DB_ws.column_dimensions["P"].width = len("Utilisateur ecriture") + 1;
for i in range(2, TEMP_WORK_5_LAST_YEARS + 2): 
    DB_ws.cell(row = i, column =16).value = "UTIL1";

DB_ws["Q1"] = "Date piece";
DB_ws.column_dimensions["Q"].width = len("Date piece") + 1;
DB_ws["R1"] = "Date comptable";
DB_ws.column_dimensions["R"].width = len("Date comptable") + 1;
DB_ws["S1"] = "Date de saisie";
DB_ws.column_dimensions["S"].width = len("Date de saisie") + 1;


DB_ws["T1"] = "Compte contre partie";
DB_ws.column_dimensions["T"].width = len("Compte contre partie") + 1;
DB_ws["U1"] = "Designation compte contre partie";
DB_ws.column_dimensions["U"].width = len("Fournisseurs - Achats de biens et prestations de services");
for i in range(2, TEMP_WORK_5_LAST_YEARS + 2): 
    DB_ws.cell(row = i, column = 20).value = "40110000";
    DB_ws.cell(row = i, column = 21).value = "Fournisseurs - Achats de biens et prestations de services";


DB_ws["V1"] = "N° Ecriture";
DB_ws.column_dimensions["V"].width = len("7000000000000") + 1;
for i in range(2, TEMP_WORK_5_LAST_YEARS + 2): 
    DB_ws.cell(row = i, column = 22).value = "70000000" + str(i-1);

DB_ws["W1"] = "Commentaire ecriture";
DB_ws.column_dimensions["W"].width = len("Commentaire ecriture") + 1;
for i in range(2, TEMP_WORK_5_LAST_YEARS + 2): 
    DB_ws.cell(row = i, column = 23).value = "Commentaire";

DB_ws["X1"] = "N° contre passation";
DB_ws.column_dimensions["X"].width = len("N° contre passation") + 1;
DB_ws["Y1"] = "Commentaire contre passation";
DB_ws.column_dimensions["Y"].width = len("Commentaire contre passation") + 1;

DB_ws["Z1"] = "Devise";
DB_ws.column_dimensions["Z"].width = len("Devise") + 1;
for i in range(2, TEMP_WORK_5_LAST_YEARS + 2): 
    DB_ws.cell(row = i, column = 26).value = "€";
DB_ws["AA1"] = "Convertion en euros";
DB_ws.column_dimensions["AA"].width = len("Devise") + 1;
DB_ws["AB1"] = "Date convertion";
DB_ws.column_dimensions["AB"].width = len("Date convertion") + 1;
DB_ws["AC1"] = "Taux convertion";
DB_ws.column_dimensions["AC"].width = len("Taux convertion") + 1;
DB_ws["AD1"] = "Source convertion";
DB_ws.column_dimensions["AD"].width = len("Source convertion") + 1;

DB_ws["AE1"] = "Societe";
DB_ws.column_dimensions["AE"].width = len("Societe") + 1;
for i in range(2, TEMP_WORK_5_LAST_YEARS + 2): 
    DB_ws.cell(row = i, column = 31).value = "1001";
DB_ws["AF1"] = "Designation societe";
DB_ws.column_dimensions["AF"].width = len("Designation societe") + 1;
for i in range(2, TEMP_WORK_5_LAST_YEARS + 2): 
    DB_ws.cell(row = i, column = 32).value = "Agenge d'interim";

DB_ws["AG1"] = "Unité de quantité";
DB_ws.column_dimensions["AG"].width = len("Unité de quantité") + 1;
for i in range(2, TEMP_WORK_5_LAST_YEARS + 2): 
    DB_ws.cell(row = i, column = 33).value = "Heures";

DB_ws["AH1"] = "Quantité";
DB_ws.column_dimensions["AH"].width = len("Quantité") + 1;
for i in range(2, TEMP_WORK_5_LAST_YEARS + 2): 
    DB_ws.cell(row = i, column = 34).value = 140;
DB_ws["AI1"] = "Taux unité de quantité";
DB_ws.column_dimensions["AI"].width = len("Taux unité de quantité") + 1;
for i in range(2, TEMP_WORK_5_LAST_YEARS + 2): 
    DB_ws.cell(row = i, column = 35).value = 14.29;

DB_ws["AJ1"] = "Code mouvement";
DB_ws.column_dimensions["AJ"].width = len("Code mouvement") + 1;
DB_ws["AK1"] = "Designation mouvement";
DB_ws.column_dimensions["AK"].width = len("Designation mouvement") + 1;



retroActiveDay = lastDayOfPreviousMounth(CurrentDay)
# Medecine de travail
for i in range(2, round(EMPLOYEES * 2.5) + 2) :
    nbOfLine = nbOfLine + 1
    DB_ws.cell(row = TEMP_WORK_5_LAST_YEARS + i, column = 1).value = "64750000";
    DB_ws.cell(row = TEMP_WORK_5_LAST_YEARS + i, column = 2).value = "Medecine du travail, pharmacie";
    DB_ws.cell(row = TEMP_WORK_5_LAST_YEARS + i, column = 3).value = 6200;
    DB_ws.cell(row = TEMP_WORK_5_LAST_YEARS + i, column = 4).value = "Usine";
    DB_ws.cell(row = TEMP_WORK_5_LAST_YEARS + i, column = 7).value = 100;
    DB_ws.cell(row = TEMP_WORK_5_LAST_YEARS + i, column = 8).value = "Facture";
    
    if NbEmployee < EMPLOYEES : 
        NbEmployee = NbEmployee + 1
    else:
        NbEmployee = 1
        # retroActiveDay = lastDayOfPreviousMounth(retroActiveDay)
    DB_ws.cell(row = TEMP_WORK_5_LAST_YEARS + i, column = 9).value = "Nom" + str(NbEmployee);
    DB_ws.cell(row = TEMP_WORK_5_LAST_YEARS + i, column = 10).value = "Prenom" + str(NbEmployee);
    DB_ws.cell(row = TEMP_WORK_5_LAST_YEARS + i, column = 11).value = 100000 + int(NbEmployee);
    DB_ws.cell(row = TEMP_WORK_5_LAST_YEARS + i, column = 15).value = "0000000000" + str( nbOfLine-1);
    DB_ws.cell(row = TEMP_WORK_5_LAST_YEARS + i, column = 16).value = "UTIL1";
    DB_ws.cell(row = TEMP_WORK_5_LAST_YEARS + i, column = 20).value = "40110000";
    DB_ws.cell(row = TEMP_WORK_5_LAST_YEARS + i, column = 21).value = "Fournisseurs - Achats de biens et prestations de services";
    DB_ws.cell(row = TEMP_WORK_5_LAST_YEARS + i, column = 23).value = "Commentaire";
    DB_ws.cell(row = TEMP_WORK_5_LAST_YEARS + i, column = 26).value = "€";
    DB_ws.cell(row = TEMP_WORK_5_LAST_YEARS + i, column = 31).value = "1001";
    DB_ws.cell(row = TEMP_WORK_5_LAST_YEARS + i, column = 32).value = "Medecine du travail";
    DB_ws.cell(row = TEMP_WORK_5_LAST_YEARS + i, column = 33).value = "Heures";
    DB_ws.cell(row = TEMP_WORK_5_LAST_YEARS + i, column = 34).value = 2;
    DB_ws.cell(row = TEMP_WORK_5_LAST_YEARS + i, column = 35).value = 40;
    if NbEmployee % round(EMPLOYEES / 12 / 2) == 0: 
        retroActiveDay = lastDayOfPreviousMounth(retroActiveDay)
    DB_ws.cell(row = TEMP_WORK_5_LAST_YEARS + i, column = 12).value = str(retroActiveDay).replace('-', '/');
    DB_ws.cell(row = TEMP_WORK_5_LAST_YEARS + i, column = 17).value = str(retroActiveDay).replace('-', '/');
    DB_ws.cell(row = TEMP_WORK_5_LAST_YEARS + i, column = 18).value = str(retroActiveDay).replace('-', '/');
    DB_ws.cell(row = TEMP_WORK_5_LAST_YEARS + i, column = 19).value = str(retroActiveDay).replace('-', '/');
    DB_ws.cell(row = TEMP_WORK_5_LAST_YEARS + i, column = 22).value = "70000000" + str( nbOfLine-1);



# Fournitures non stockables (eau, énergie) - Eau 60610001
retroActiveDay = CurrentDay

for i in range(2, 12*5 + 2) :
    nbOfLine = nbOfLine + 1
    DB_ws.cell(row = nbOfLine, column = 1).value = "60610001";
    DB_ws.cell(row = nbOfLine, column = 2).value = "Fournitures non stockables (eau, énergie) - Eau";
    DB_ws.cell(row = nbOfLine, column = 3).value = 6200;
    DB_ws.cell(row = nbOfLine, column = 4).value = "Usine";
    DB_ws.cell(row = nbOfLine, column = 7).value = 2000;
    DB_ws.cell(row = nbOfLine, column = 8).value = "Facture";
    retroActiveDay = lastDayOfPreviousMounth(retroActiveDay)
    DB_ws.cell(row = nbOfLine, column = 12).value = str(retroActiveDay).replace('-', '/');
    DB_ws.cell(row = nbOfLine, column = 15).value = "0000000000" + str( nbOfLine-1);
    DB_ws.cell(row = nbOfLine, column = 16).value = "UTIL1";
    DB_ws.cell(row = nbOfLine, column = 17).value = str(retroActiveDay).replace('-', '/');
    DB_ws.cell(row = nbOfLine, column = 18).value = str(retroActiveDay).replace('-', '/');
    DB_ws.cell(row = nbOfLine, column = 19).value = str(retroActiveDay).replace('-', '/');
    DB_ws.cell(row = nbOfLine, column = 20).value = "40110000";
    DB_ws.cell(row = nbOfLine, column = 21).value = "Fournisseurs - Achats de biens et prestations de services";
    DB_ws.cell(row = nbOfLine, column = 22).value = "70000000" + str( nbOfLine-1);
    DB_ws.cell(row = nbOfLine, column = 23).value = "Commentaire";
    DB_ws.cell(row = nbOfLine, column = 26).value = "€";
    DB_ws.cell(row = nbOfLine, column = 31).value = "1001";
    DB_ws.cell(row = nbOfLine, column = 32).value = "Eau";
    DB_ws.cell(row = nbOfLine, column = 33).value = "Litre";
    DB_ws.cell(row = nbOfLine, column = 34).value = 5000;
    DB_ws.cell(row = nbOfLine, column = 35).value = 1;



# Fournitures non stockables (eau, énergie) - Gaz 60610002
retroActiveDay = CurrentDay

for i in range(2, 12*5 + 2) :
    nbOfLine = nbOfLine + 1
    retroActiveDay = lastDayOfPreviousMounth(retroActiveDay)
    DB_ws.cell(row = nbOfLine, column = 1).value = "60610002";
    DB_ws.cell(row = nbOfLine, column = 2).value = "Fournitures non stockables (eau, énergie) - Gaz";
    DB_ws.cell(row = nbOfLine, column = 3).value = 6200;
    DB_ws.cell(row = nbOfLine, column = 4).value = "Usine";
    DB_ws.cell(row = nbOfLine, column = 7).value = GAZ_CONSUMPTION * GAZ_COST[2018 - int(str(retroActiveDay).split("-")[0])];
    DB_ws.cell(row = nbOfLine, column = 8).value = "Facture";
    DB_ws.cell(row = nbOfLine, column = 12).value = str(retroActiveDay).replace('-', '/');
    DB_ws.cell(row = nbOfLine, column = 15).value = "0000000000" + str( nbOfLine-1);
    DB_ws.cell(row = nbOfLine, column = 16).value = "UTIL1";
    DB_ws.cell(row = nbOfLine, column = 17).value = str(retroActiveDay).replace('-', '/');
    DB_ws.cell(row = nbOfLine, column = 18).value = str(retroActiveDay).replace('-', '/');
    DB_ws.cell(row = nbOfLine, column = 19).value = str(retroActiveDay).replace('-', '/');
    DB_ws.cell(row = nbOfLine, column = 20).value = "40110000";
    DB_ws.cell(row = nbOfLine, column = 21).value = "Fournisseurs - Achats de biens et prestations de services";
    DB_ws.cell(row = nbOfLine, column = 22).value = "70000000" + str( nbOfLine-1);
    DB_ws.cell(row = nbOfLine, column = 23).value = "Commentaire";
    DB_ws.cell(row = nbOfLine, column = 26).value = "€";
    DB_ws.cell(row = nbOfLine, column = 31).value = "1001";
    DB_ws.cell(row = nbOfLine, column = 32).value = "gaz";
    DB_ws.cell(row = nbOfLine, column = 33).value = "KwH";
    DB_ws.cell(row = nbOfLine, column = 34).value = 100000;
    DB_ws.cell(row = nbOfLine, column = 35).value = GAZ_COST[2018 - int(str(retroActiveDay).split("-")[0])];


# Fournitures non stockables (eau, énergie) - Electricité 60610003
retroActiveDay = CurrentDay

for i in range(2, 12*5 + 2) :
    nbOfLine = nbOfLine + 1
    retroActiveDay = lastDayOfPreviousMounth(retroActiveDay)
    DB_ws.cell(row = nbOfLine, column = 1).value = "60610003";
    DB_ws.cell(row = nbOfLine, column = 2).value = "Fournitures non stockables (eau, énergie) - Electricité";
    DB_ws.cell(row = nbOfLine, column = 3).value = 6200;
    DB_ws.cell(row = nbOfLine, column = 4).value = "Usine";
    DB_ws.cell(row = nbOfLine, column = 7).value = ELECTRICITY_CONSUMPTION * ELECTRICITY_COST[2018 - int(str(retroActiveDay).split("-")[0])];
    DB_ws.cell(row = nbOfLine, column = 8).value = "Facture";
    DB_ws.cell(row = nbOfLine, column = 12).value = str(retroActiveDay).replace('-', '/');
    DB_ws.cell(row = nbOfLine, column = 15).value = "0000000000" + str( nbOfLine-1);
    DB_ws.cell(row = nbOfLine, column = 16).value = "UTIL1";
    DB_ws.cell(row = nbOfLine, column = 17).value = str(retroActiveDay).replace('-', '/');
    DB_ws.cell(row = nbOfLine, column = 18).value = str(retroActiveDay).replace('-', '/');
    DB_ws.cell(row = nbOfLine, column = 19).value = str(retroActiveDay).replace('-', '/');
    DB_ws.cell(row = nbOfLine, column = 20).value = "40110000";
    DB_ws.cell(row = nbOfLine, column = 21).value = "Fournisseurs - Achats de biens et prestations de services";
    DB_ws.cell(row = nbOfLine, column = 22).value = "70000000" + str( nbOfLine-1);
    DB_ws.cell(row = nbOfLine, column = 23).value = "Commentaire";
    DB_ws.cell(row = nbOfLine, column = 26).value = "€";
    DB_ws.cell(row = nbOfLine, column = 31).value = "1001";
    DB_ws.cell(row = nbOfLine, column = 32).value = "Electricité";
    DB_ws.cell(row = nbOfLine, column = 33).value = "KwH";
    DB_ws.cell(row = nbOfLine, column = 34).value = 50000;
    DB_ws.cell(row = nbOfLine, column = 35).value = ELECTRICITY_COST[2018 - int(str(retroActiveDay).split("-")[0])];


# Sous-traitance générale - Cantine 61100000
retroActiveDay = CurrentDay

for i in range(2, 12*5 + 2) :
    nbOfLine = nbOfLine + 1
    DB_ws.cell(row = nbOfLine, column = 1).value = "61100000";
    DB_ws.cell(row = nbOfLine, column = 2).value = "Sous-traitance générale - Cantine";
    DB_ws.cell(row = nbOfLine, column = 3).value = 6200;
    DB_ws.cell(row = nbOfLine, column = 4).value = "Usine";
    DB_ws.cell(row = nbOfLine, column = 7).value = 2100;
    DB_ws.cell(row = nbOfLine, column = 8).value = "Facture";
    retroActiveDay = lastDayOfPreviousMounth(retroActiveDay)
    DB_ws.cell(row = nbOfLine, column = 12).value = str(retroActiveDay).replace('-', '/');
    DB_ws.cell(row = nbOfLine, column = 15).value = "0000000000" + str( nbOfLine-1);
    DB_ws.cell(row = nbOfLine, column = 16).value = "UTIL1";
    DB_ws.cell(row = nbOfLine, column = 17).value = str(retroActiveDay).replace('-', '/');
    DB_ws.cell(row = nbOfLine, column = 18).value = str(retroActiveDay).replace('-', '/');
    DB_ws.cell(row = nbOfLine, column = 19).value = str(retroActiveDay).replace('-', '/');
    DB_ws.cell(row = nbOfLine, column = 20).value = "40110000";
    DB_ws.cell(row = nbOfLine, column = 21).value = "Fournisseurs - Achats de biens et prestations de services";
    DB_ws.cell(row = nbOfLine, column = 22).value = "70000000" + str( nbOfLine-1);
    DB_ws.cell(row = nbOfLine, column = 23).value = "Commentaire";
    DB_ws.cell(row = nbOfLine, column = 26).value = "€";
    DB_ws.cell(row = nbOfLine, column = 31).value = "1001";
    DB_ws.cell(row = nbOfLine, column = 32).value = "Cantine";
    DB_ws.cell(row = nbOfLine, column = 33).value = "Heure";
    DB_ws.cell(row = nbOfLine, column = 34).value = 140;
    DB_ws.cell(row = nbOfLine, column = 35).value = 15;


# Entretien et réparations sur biens immobiliers 61520000
retroActiveDay = CurrentDay

for i in range(2, 12*5 + 2) :
    nbOfLine = nbOfLine + 1
    DB_ws.cell(row = nbOfLine, column = 1).value = "61520000";
    DB_ws.cell(row = nbOfLine, column = 2).value = "Entretien et réparations sur biens immobiliers";
    DB_ws.cell(row = nbOfLine, column = 3).value = 6200;
    DB_ws.cell(row = nbOfLine, column = 4).value = "Usine";
    DB_ws.cell(row = nbOfLine, column = 7).value = 10000;
    DB_ws.cell(row = nbOfLine, column = 8).value = "Facture";
    retroActiveDay = lastDayOfPreviousMounth(retroActiveDay)
    DB_ws.cell(row = nbOfLine, column = 12).value = str(retroActiveDay).replace('-', '/');
    DB_ws.cell(row = nbOfLine, column = 15).value = "0000000000" + str( nbOfLine-1);
    DB_ws.cell(row = nbOfLine, column = 16).value = "UTIL1";
    DB_ws.cell(row = nbOfLine, column = 17).value = str(retroActiveDay).replace('-', '/');
    DB_ws.cell(row = nbOfLine, column = 18).value = str(retroActiveDay).replace('-', '/');
    DB_ws.cell(row = nbOfLine, column = 19).value = str(retroActiveDay).replace('-', '/');
    DB_ws.cell(row = nbOfLine, column = 20).value = "40110000";
    DB_ws.cell(row = nbOfLine, column = 21).value = "Fournisseurs - Achats de biens et prestations de services";
    DB_ws.cell(row = nbOfLine, column = 22).value = "70000000" + str( nbOfLine-1);
    DB_ws.cell(row = nbOfLine, column = 23).value = "Commentaire";
    DB_ws.cell(row = nbOfLine, column = 26).value = "€";
    DB_ws.cell(row = nbOfLine, column = 31).value = "1001";
    DB_ws.cell(row = nbOfLine, column = 32).value = "Nettoyage des locaux";
    DB_ws.cell(row = nbOfLine, column = 33).value = "Heure";
    DB_ws.cell(row = nbOfLine, column = 34).value = 140;
    DB_ws.cell(row = nbOfLine, column = 35).value = 15;


DB_ws["AM1"] = "Annee";
DB_ws["AN1"] = "Mois";
DB_ws["AO1"] = "Jour";
# Commencer à partir de la deuxième ligne si la première ligne contient des en-têtes
for row in DB_ws.iter_rows(min_row=2, min_col=18, max_col=18):  # Colonne R (18e colonne)
    for cell in row:
        # Extraire la valeur de la cellule (format YYYY-MM-DD)
        if cell.value:
            date_str = str(cell.value)
            
            try:
                # Splitter la date pour obtenir année, mois, jour
                year, month, day = date_str.split('/')

                # Insérer dans les colonnes AM (38), AN (39), AO (40)
                DB_ws.cell(row=cell.row, column=39, value=int(year))  # Colonne AM
                DB_ws.cell(row=cell.row, column=40, value=int(month))  # Colonne AN
                DB_ws.cell(row=cell.row, column=41, value=int(day))  # Colonne AO

            except ValueError:
                print(f"Erreur lors du traitement de la date : {date_str}")

# Parcourir les lignes, en supposant que les clés se trouvent dans la colonne B (colonne 2)
DB_ws["AL1"] = "Abreviation";
for row in DB_ws.iter_rows(min_row=2, max_row=DB_ws.max_row, min_col=2, max_col=2):
    key_cell = row[0]  # Cellule en colonne B
    key_value = key_cell.value

    if key_value in dictionnaire:
        # Affecter la valeur dans la colonne AL (colonne 38)
        DB_ws.cell(row=key_cell.row, column=38, value=dictionnaire[key_value])


DB_ws["AP1"] = "Mois Texte"
for i in range(2, nbOfLine + 1):
    date_in_AN = DB_ws.cell(row=i, column=40).value
    DB_ws.cell(row=i, column=42).value = text_month[str(date_in_AN)]  # Obtenir le mois en texte

from random import sample

num_rows_to_delete = randint(round(nbOfLine * 0.1), round(nbOfLine * 0.2))

rows_to_delete = sample(range(2, nbOfLine+1), num_rows_to_delete)
rows_to_delete.sort(reverse=True)

for row in rows_to_delete:
    DB_ws.delete_rows(row)
    print(" ", row)



def cellToNums(cell) :
    pass

def numsToCell(row, col) :
    pass


def rangeToNums(range) : 
    pass

def numsToRange(startRow, startCol, endRow, endCol) : 
    pass



income_statement_DB_file.save("./income_statement_DBGenerator.xlsx");