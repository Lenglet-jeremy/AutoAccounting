from openpyxl.workbook import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import load_workbook

def range_to_numbers(Range) :
    Col_start = ""
    Row_start = 0

    Col_end = ""
    Row_end = 0

    if ":" in Range :

        Range_splitted = Range.split(":") 

        for i, c in enumerate(Range_splitted[0]) : 
            try : 
                Row_start = int(Range_splitted[0][i:])
                break
            except ValueError : 
                Col_start = Col_start + c

        for i, c in enumerate(Range_splitted[1]) : 
            try : 
                Row_end = int(Range_splitted[1][i:])
                break
            except ValueError : 
                Col_end = Col_end + c

        Col_start = col_string_to_num(Col_start)
        Col_end = col_string_to_num(Col_end)

        Tab = [Col_start, Row_start, Col_end, Row_end]
        return Tab
    
    else : 
        for i, c in enumerate(Range) : 
            try : 
                Row_start = int(Range[i:])
                break
            except ValueError : 
                Col_start = Col_start + c
        
        Col_start = col_string_to_num(Col_start)
        Tab = [Col_start, Row_start]
        return Tab


def numbers_to_range(col_start, row_start, col_end = None, row_end = None) : 

    if col_end == None and row_end == None  : 

        start_col = str(get_column_letter(col_start))
        start_row = str(row_start)

        return start_col + start_row
    else :
        start_col = str(get_column_letter(col_start))
        start_row = str(row_start)
        end_col = str(get_column_letter(col_end))
        end_row = str(row_end)
        
        return start_col + start_row + ":" + end_col + end_row

def horizontal_range_travel(Range) : 
        Tab = range_to_numbers(Range)
        Col_start = Tab[0]
        Row_start = Tab[1]

        Col_end = Tab[2]
        Row_end = Tab[3]

        New_range = numbers_to_range(col_start = Col_start + 1, row_start = Row_start, col_end = Col_end, row_end = Row_end)
        Index = New_range.split(":")
        current_cell = Index[0]
        Index_end = Index[1]

        New_cell = current_cell
        return New_cell

def vertical_range_travel(Range) : 
        Tab = range_to_numbers(Range)
        Col_start = Tab[0]
        Row_start = Tab[1]

        Col_end = Tab[2]
        Row_end = Tab[3]

        New_range = numbers_to_range(col_start = Col_start, row_start = Row_start + 1, col_end = Col_end, row_end = Row_end)
        Index = New_range.split(":")
        current_cell = Index[0]
        Index_end = Index[1]

        New_cell = current_cell
        return New_cell

def Empty_styling(Worksheet, Range) :
    Empty_cell_font = Font(name = "Arial", size = 8, color = "FFFFFF")
    Empty_cell_pattern = PatternFill(fill_type = "lightGray",fgColor= "000000", bgColor="000080")
    Empty_cell_alignment = Alignment(horizontal = "center", vertical = "center")
    Empty_cell_border = Border( top = Side(border_style = "hair", color = "000000"),
                                right = Side(border_style = "hair", color = "000000"),
                                bottom = Side(border_style = "hair", color = "000000"),
                                left = Side(border_style = "hair", color = "000000"))
    Tab = range_to_numbers(Range)

    col_start = Tab[0]
    row_start = Tab[1]
    
    col_end = Tab[2] 
    row_end = Tab[3]

    for col in range(col_start, col_end + 1) : 
        for row in range(row_start, row_end + 1) : 
            Worksheet.cell(row = row, column = col).font = Empty_cell_font
            Worksheet.cell(row = row, column = col).fill = Empty_cell_pattern
            Worksheet.cell(row = row, column = col).border = Empty_cell_border
            Worksheet.cell(row = row, column = col).alignment = Empty_cell_alignment

def External_formatting(Worksheet, Range, Type_border) :

    tab_range = range_to_numbers(Range)

    
    for col in range(tab_range[0], tab_range[2] + 1) : 
        for row in range(tab_range[1], tab_range[3] + 1) : 
    


            Top_border = Border(    top = Side(border_style = Type_border, color = "000000"),
                                    right = Side(border_style = Worksheet.cell(row = row, column = col).border.right.style, color = "000000"),
                                    bottom = Side(border_style = Worksheet.cell(row = row, column = col).border.bottom.style, color = "000000"),
                                    left = Side(border_style = Worksheet.cell(row = row, column = col).border.left.style, color = "000000"))
    
            Right_border = Border(  top = Side(border_style = Worksheet.cell(row = row, column = col).border.top.style, color = "000000"),
                                    right = Side(border_style = Type_border, color = "000000"),
                                    bottom = Side(border_style = Worksheet.cell(row = row, column = col).border.bottom.style, color = "000000"),
                                    left = Side(border_style = Worksheet.cell(row = row, column = col).border.left.style, color = "000000"))
    
            Bottom_border = Border( top = Side(border_style = Worksheet.cell(row = row, column = col).border.top.style, color = "000000"),
                                    right = Side(border_style = Worksheet.cell(row = row, column = col).border.right.style, color = "000000"),
                                    bottom = Side(border_style = Type_border, color = "000000"),
                                    left = Side(border_style = Worksheet.cell(row = row, column = col).border.left.style, color = "000000"))
    
            Left_border = Border(   top = Side(border_style = Worksheet.cell(row = row, column = col).border.top.style, color = "000000"),
                                    right = Side(border_style = Worksheet.cell(row = row, column = col).border.right.style, color = "000000"),
                                    bottom = Side(border_style = Worksheet.cell(row = row, column = col).border.bottom.style, color = "000000"),
                                    left = Side(border_style = Type_border, color = "000000"))

            Top_left_border = Border(   top = Side(border_style = Type_border, color = "000000"),
                                        right = Side(border_style = Worksheet.cell(row = row, column = col).border.right.style, color = "000000"),
                                        bottom = Side(border_style = Worksheet.cell(row = row, column = col).border.bottom.style, color = "000000"),
                                        left = Side(border_style = Type_border, color = "000000"))

            Top_right_border = Border(  top = Side(border_style = Type_border, color = "000000"),
                                        right = Side(border_style = Type_border, color = "000000"),
                                        bottom = Side(border_style = Worksheet.cell(row = row, column = col).border.bottom.style, color = "000000"),
                                        left = Side(border_style = Worksheet.cell(row = row, column = col).border.left.style, color = "000000"))

            Bottom_left_border = Border(top = Side(border_style = Worksheet.cell(row = row, column = col).border.top.style, color = "000000"),
                                        right = Side(border_style = Worksheet.cell(row = row, column = col).border.right.style, color = "000000"),
                                        bottom = Side(border_style = Type_border, color = "000000"),
                                        left = Side(border_style = Type_border, color = "000000"))

            Bottom_right_border = Border(   top = Side(border_style = Worksheet.cell(row = row, column = col).border.top.style, color = "000000"),
                                            right = Side(border_style = Type_border, color = "000000"),
                                            bottom = Side(border_style = Type_border, color = "000000"),
                                            left = Side(border_style = Worksheet.cell(row = row, column = col).border.left.style, color = "000000"))
            


            
            if col == tab_range[0] and row == tab_range[1] : 
                Worksheet.cell(row, col).border = Top_left_border

            elif col == tab_range[2] and row == tab_range[1] : 
                Worksheet.cell(row, col).border = Top_right_border

            elif col == tab_range[0] and row == tab_range[3] :
                Worksheet.cell(row, col).border = Bottom_left_border

            elif col == tab_range[2] and row == tab_range[3] :
                Worksheet.cell(row, col).border = Bottom_right_border
            
            elif row == tab_range[1] :
                Worksheet.cell(row, col).border = Top_border

            elif col == tab_range[2] :
                Worksheet.cell(row, col).border = Right_border

            elif row == tab_range[3] :
                Worksheet.cell(row, col).border = Bottom_border
            
            elif col == tab_range[0] : 
                Worksheet.cell(row, col).border = Left_border

def Internal_formating(Worksheet, Range, Type_border) :
    tab_range = range_to_numbers(Range)


    for col in range(tab_range[0], tab_range[2] + 1) : 
        for row in range(tab_range[1], tab_range[3] + 1) : 
    
            Borders = Border(   top = Side(border_style = Type_border, color = "000000"),
                                right = Side(border_style = Type_border, color = "000000"),
                                bottom = Side(border_style = Type_border, color = "000000"),
                                left = Side(border_style = Type_border, color = "000000"))
            
            Worksheet.cell(row, col).border = Borders
    

def Top_border_formating(Worksheet, Range, Type_border) :
    
    tab_range = range_to_numbers(Range)


    for col in range(tab_range[0], tab_range[2] + 1) : 
        for row in range(tab_range[1], tab_range[3] + 1) : 
            Top_border = Border(top = Side(border_style = Type_border, color = "000000"),
                                right = Side(border_style = Worksheet.cell(row = row, column = col).border.right.style, color = "000000"),
                                bottom = Side(border_style = Worksheet.cell(row = row, column = col).border.bottom.style, color = "000000"),
                                left = Side(border_style = Worksheet.cell(row = row, column = col).border.left.style, color = "000000"))
            
            Worksheet.cell(row, col).border = Top_border

def Right_border_formating(Worksheet, Range, Type_border) :
    
    tab_range = range_to_numbers(Range)


    for col in range(tab_range[0], tab_range[2] + 1) : 
        for row in range(tab_range[1], tab_range[3] + 1) : 
            Right_border = Border(  top = Side(border_style = Worksheet.cell(row = row, column = col).border.top.style, color = "000000"),
                                    right = Side(border_style = Type_border, color = "000000"),
                                    bottom = Side(border_style = Worksheet.cell(row = row, column = col).border.bottom.style, color = "000000"),
                                    left = Side(border_style = Worksheet.cell(row = row, column = col).border.left.style, color = "000000"))
            
            Worksheet.cell(row, col).border = Right_border
            
def Bottom_border_formating(Worksheet, Range, Type_border) :
    
    tab_range = range_to_numbers(Range)


    for col in range(tab_range[0], tab_range[2] + 1) : 
        for row in range(tab_range[1], tab_range[3] + 1) : 
            
            Bottom_border = Border( top = Side(border_style = Worksheet.cell(row = row, column = col).border.top.style, color = "000000"),
                                    right = Side(border_style = Worksheet.cell(row = row, column = col).border.right.style, color = "000000"),
                                    bottom = Side(border_style = Type_border, color = "000000"),
                                    left = Side(border_style = Worksheet.cell(row = row, column = col).border.left.style, color = "000000"))
            
            Worksheet.cell(row, col).border = Bottom_border
            
def Left_border_formating(Worksheet, Range, Type_border) :
    
    tab_range = range_to_numbers(Range)


    for col in range(tab_range[0], tab_range[2] + 1) : 
        for row in range(tab_range[1], tab_range[3] + 1) : 
            
            Left_border = Border(   top = Side(border_style = Worksheet.cell(row = row, column = col).border.top.style, color = "000000"),
                                    right = Side(border_style = Worksheet.cell(row = row, column = col).border.right.style, color = "000000"),
                                    bottom = Side(border_style = Worksheet.cell(row = row, column = col).border.bottom.style, color = "000000"),
                                    left = Side(border_style = Type_border, color = "000000"))
            
            Worksheet.cell(row, col).border = Left_border


def col_string_to_num(s):
    # Source from https://stackoverflow.com/a/63013258
    n = ord(s[-1]) - 64 # To set 'A' to 1
    if s[:-1]: #until not empty
        return 26 * (col_string_to_num(s[:-1])) + n
    else:
        return n
    
# Function to convert a given number to an Excel column
def getColumnName(n):
    #Source : https://www.techiedelight.com/convert-given-number-corresponding-excel-column-name/
    # initialize output string as empty
    result = ''
 
    while n > 0:
 
        # find the index of the next letter and concatenate the letter
        # to the solution
 
        # here index 0 corresponds to 'A', and 25 corresponds to 'Z'
        index = (n - 1) % 26
        result += chr(index + ord('A'))
        n = (n - 1) // 26
 
    return result[::-1]

PROVISIONS_FILE_PATH = "E:/A0_Boulot Perso-Pro/0_Projects/Automatisation compta/Updating Files/Provision file/Provisions.xlsx"
Provisions_file = Workbook()

def Reference_worksheet() :
    #To do :
        #Implement holydays

    reference_ws = Provisions_file.active

    reference_ws.title = "Référence"

    #Variable's initialization
    months = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet", "Aout", "Septembre", "Octobre", "Novembre", "Decembre"]
    days = ["L", "M", "M", "J", "V", "S", "D"]
    days_position = 6
    VALUES_DAYS_CALENDAR = {"cell index" : "value cell"}


    reference_ws.column_dimensions["A"].width = 20/7
    reference_ws.column_dimensions["B"].width = 20/7

    #Formatting collumns of calendar
    for col in range(3,38 + 1) :
        if col % 3 == 0 : 
            reference_ws.column_dimensions[getColumnName(col)].width = 25/7

        if col % 3 == 1 : 
            reference_ws.column_dimensions[getColumnName(col)].width = 20/7

        if col % 3 == 2 : 
            reference_ws.column_dimensions[getColumnName(col)].width = 70/7

    reference_ws.column_dimensions["AM"].width = 20/7

    reference_ws.column_dimensions["AN"].width = 120/7
    reference_ws.column_dimensions["AO"].width = 70/7
    reference_ws.column_dimensions["AP"].width = 70/7
    reference_ws.column_dimensions["AQ"].width = 70/7
    reference_ws.column_dimensions["AR"].width = 70/7

    reference_ws.column_dimensions["AS"].width = 80/7
    
    reference_ws.column_dimensions["AT"].width = 120/7
    reference_ws.column_dimensions["AU"].width = 120/7
    
    reference_ws.column_dimensions["AX"].width = 25/7
    reference_ws.column_dimensions["AY"].width = 25/7
    reference_ws.column_dimensions["AZ"].width = 25/7
    
    reference_ws.column_dimensions["BA"].width = 90/7
    reference_ws.column_dimensions["BB"].width = 90/7
    
    reference_ws.column_dimensions["BC"].width = 25/7
    reference_ws.column_dimensions["BD"].width = 25/7
    reference_ws.column_dimensions["BE"].width = 25/7
    
    reference_ws.column_dimensions["BF"].width = 80/7
    reference_ws.column_dimensions["BG"].width = 370/7
    reference_ws.column_dimensions["BH"].width = 80/7
    
    reference_ws.column_dimensions["BI"].width = 25/7
    reference_ws.column_dimensions["BJ"].width = 25/7
    reference_ws.column_dimensions["BK"].width = 25/7

    reference_ws.column_dimensions["BL"].width = 100/7
    reference_ws.column_dimensions["BM"].width = 160/7

    #Formatting head of calendar
    reference_ws.merge_cells('C1:AL1')
    reference_ws['C1'] = "Calendrier 2023"
    reference_ws['C1'].alignment = Alignment("center", "center")
    reference_ws['C1'].font = Font(name = "Arial" , sz = 24, color = "000040")

    Range_splitted = 'C1:AM1'.split(":")
    current_cell = Range_splitted[0]
    while current_cell != Range_splitted[1] : 
        reference_ws[current_cell].border = Border( top = Side(border_style = "thin", color= "000000"),
                                                    right = Side(border_style = "thin", color = "000000"),
                                                    bottom = Side(border_style = "thin", color = "000000"),
                                                    left = Side(border_style = "thin", color = "000000"))
        current_cell = horizontal_range_travel(current_cell + ":" + Range_splitted[1])
    
    #Draw months
    for month in range(0,12) : 
        
        reference_ws.merge_cells(numbers_to_range(col_start = month * 3 + 3, row_start = 2, col_end = month * 3 + 5, row_end = 2))
        reference_ws.cell(column = month * 3 + 3 , row = 2).value = months[month]
        reference_ws.cell(column = month * 3 + 3 , row = 2).fill = PatternFill(start_color = "000040", end_color = "000040", fill_type = "solid")
        reference_ws.cell(column = month * 3 + 3 , row = 2).font = Font(name = "Arial", sz = 8, color = "FFFFFF")
        reference_ws.cell(column = month * 3 + 3 , row = 2).alignment = Alignment("center", "center")
        
        reference_ws.cell(column = month * 3 + 3 , row = 2).border = Border(top = Side(border_style = "hair", color= "000000"),
                                                                            right = Side(border_style = "hair", color = "000000"),
                                                                            bottom = Side(border_style = "hair", color = "000000"),
                                                                            left = Side(border_style = "hair", color = "000000"))
        reference_ws.cell(column = month * 3 + 4 , row = 2).border = Border(top = Side(border_style = "hair", color= "000000"),
                                                                            right = Side(border_style = "hair", color = "000000"),
                                                                            bottom = Side(border_style = "hair", color = "000000"),
                                                                            left = Side(border_style = "hair", color = "000000"))
        reference_ws.cell(column = month * 3 + 5 , row = 2).border = Border(top = Side(border_style = "hair", color= "000000"),
                                                                            right = Side(border_style = "hair", color = "000000"),
                                                                            bottom = Side(border_style = "hair", color = "000000"),
                                                                            left = Side(border_style = "hair", color = "000000"))
        

    #Draw day's numbers
        for day in range(0, 31) :


            if month == 1 and day >= 28 : 
                break

            if month == 3 and day >= 30 : 
                break

            if month == 5 and day >= 30 : 
                break

            if month == 8 and day >= 30 : 
                break

            if month == 10 and day >= 30 : 
                break

            reference_ws.cell(column = month * 3 + 3, row = day + 3).value = day + 1
            reference_ws.cell(column = month * 3 + 3, row = day + 3).fill = PatternFill(start_color = "000040", end_color = "000040", fill_type = "solid")
            reference_ws.cell(column = month * 3 + 3, row = day + 3).font = Font(name = "Arial", sz = 8, color = "FFFFFF")
            reference_ws.cell(column = month * 3 + 3, row = day + 3).alignment = Alignment("center", "center")
            reference_ws.cell(column = month * 3 + 3, row = day + 3).border = Border(   top = Side(border_style = "hair", color= "000000"),
                                                                                        right = Side(border_style = "hair", color = "000000"),
                                                                                        bottom = Side(border_style = "hair", color = "000000"),
                                                                                        left = Side(border_style = "hair", color = "000000"))

            #Draw sunday
            if days_position == 6 : 

                reference_ws.cell(column = month * 3 + 4, row = day + 3).value =  days[days_position]
                reference_ws.cell(column = month * 3 + 4, row = day + 3).fill = PatternFill(start_color = "000080", end_color = "000080", fill_type = "solid")
                reference_ws.cell(column = month * 3 + 4, row = day + 3).font = Font(name = "Arial", sz = 8, color = "FFFFFF")
                reference_ws.cell(column = month * 3 + 4, row = day + 3).alignment = Alignment("center", "center")
                reference_ws.cell(column = month * 3 + 4, row = day + 3).border = Border(   top = Side(border_style = "hair", color= "000000"),
                                                                                            right = Side(border_style = "hair", color = "000000"),
                                                                                            bottom = Side(border_style = "hair", color = "000000"),
                                                                                            left = Side(border_style = "hair", color = "000000"))

                reference_ws.cell(column = month * 3 + 5, row = day + 3).fill = PatternFill(start_color = "000080", end_color = "000080", fill_type = "solid")
                reference_ws.cell(column = month * 3 + 5, row = day + 3).font = Font(name = "Arial", sz = 8, color = "FFFFFF")
                reference_ws.cell(column = month * 3 + 5, row = day + 3).alignment = Alignment("center", "center")
                reference_ws.cell(column = month * 3 + 5, row = day + 3).border = Border(   top = Side(border_style = "hair", color= "000000"),
                                                                                            right = Side(border_style = "hair", color = "000000"),
                                                                                            bottom = Side(border_style = "hair", color = "000000"),
                                                                                            left = Side(border_style = "hair", color = "000000"))

                days_position = 0

            #Draw days of week
            else : 
                reference_ws.cell(column = month * 3 + 4, row = day + 3).value =  days[days_position]
                reference_ws.cell(column = month * 3 + 4, row = day + 3).alignment = Alignment("center", "center")
                reference_ws.cell(column = month * 3 + 4, row = day + 3).font = Font(name = "Arial", sz = 8, color = "000000")
                reference_ws.cell(column = month * 3 + 4, row = day + 3).border = Border(   top = Side(border_style = "hair", color= "000000"),
                                                                                            right = Side(border_style = "hair", color = "000000"),
                                                                                            bottom = Side(border_style = "hair", color = "000000"),
                                                                                            left = Side(border_style = "hair", color = "000000"))
                reference_ws.cell(column = month * 3 + 5, row= day + 3).border = Border(    top = Side(border_style = "hair", color= "000000"),
                                                                                            right = Side(border_style = "hair", color = "000000"),
                                                                                            bottom = Side(border_style = "hair", color = "000000"),
                                                                                            left = Side(border_style = "hair", color = "000000"))
                days_position = days_position + 1

            #Sum of invoices received per day
            #Replace these values ​​once the database is created
            reference_ws.cell(column = month * 3 + 5, row = day + 3).value =  VALUES_DAYS_CALENDAR["cell index"]

            
        #Unlikely days
        Empty_styling(Worksheet = reference_ws, Range = "F31:H33")
        Empty_styling(Worksheet = reference_ws, Range = "L33:N33")
        Empty_styling(Worksheet = reference_ws, Range = "R33:T33")
        Empty_styling(Worksheet = reference_ws, Range = "AA33:AC33")
        Empty_styling(Worksheet = reference_ws, Range = "AG33:AI33")

    #Border around calendar
    External_formatting(Worksheet = reference_ws, Range = "C2:AL33", Type_border = "thin")

    
    #Pieces board
    reference_ws["AT2"].value = "Types piéces"
    reference_ws["AT2"].fill = PatternFill(start_color = "000040", end_color = "000040", fill_type = "solid")
    reference_ws["AT2"].font = Font(name = "Arial", sz = 8, color = "FFFFFF")
    reference_ws["AT3"].value = "*"
    reference_ws["AT4"].value = "Facture"
    reference_ws["AT5"].value = "Avoir"
    reference_ws["AT6"].value = "Provisions"
    reference_ws["AT7"].value = "Contre passassion"

    reference_ws["AU2"].value = "Catégories Piéces"
    reference_ws["AU2"].fill = PatternFill(start_color = "000040", end_color = "000040", fill_type = "solid")
    reference_ws["AU2"].font = Font(name = "Arial", sz = 8, color = "FFFFFF")
    reference_ws["AU3"].value = "*"
    reference_ws["AU4"].value = "Interim"
    reference_ws["AU5"].value = "Pole Sante"
    reference_ws["AU6"].value = "Eau"
    reference_ws["AU7"].value = "Gaz"
    reference_ws["AU8"].value = "Electricite"
    reference_ws["AU9"].value = "Cantine"
    reference_ws["AU10"].value = "Nettoyage"
    reference_ws["AU11"].value = "Transport"

    #Border around pieces board
    Internal_formating(Worksheet = reference_ws, Range = "AT2:AU11", Type_border = "hair")
    Bottom_border_formating(Worksheet = reference_ws, Range = "AT2:AU2", Type_border = "dotted")
    External_formatting(Worksheet = reference_ws, Range = "AT2:AU11", Type_border = "thin")

    #Color code board
    reference_ws["AN2"].value = "Désignation"
    reference_ws["AN2"].fill = PatternFill(start_color = "000040", end_color = "000040", fill_type = "solid")
    reference_ws["AN2"].font = Font(name = "Arial", sz = 8, color = "FFFFFF")
    reference_ws["AN2"].alignment = Alignment("center", "center")

    reference_ws["AO2"].value = "R"
    reference_ws["AO2"].fill = PatternFill(start_color = "000040", end_color = "000040", fill_type = "solid")
    reference_ws["AO2"].font = Font(name = "Arial", sz = 8, color = "FFFFFF")
    reference_ws["AO2"].alignment = Alignment("center", "center")

    reference_ws["AP2"].value = "V"
    reference_ws["AP2"].fill = PatternFill(start_color = "000040", end_color = "000040", fill_type = "solid")
    reference_ws["AP2"].font = Font(name = "Arial", sz = 8, color = "FFFFFF")
    reference_ws["AP2"].alignment = Alignment("center", "center")

    reference_ws["AQ2"].value = "B"
    reference_ws["AQ2"].fill = PatternFill(start_color = "000040", end_color = "000040", fill_type = "solid")
    reference_ws["AQ2"].font = Font(name = "Arial", sz = 8, color = "FFFFFF")
    reference_ws["AQ2"].alignment = Alignment("center", "center")

    reference_ws["AR2"].value = "HEX"
    reference_ws["AR2"].fill = PatternFill(start_color = "000040", end_color = "000040", fill_type = "solid")
    reference_ws["AR2"].font = Font(name = "Arial", sz = 8, color = "FFFFFF")
    reference_ws["AR2"].alignment = Alignment("center", "center")

    reference_ws["AN3"].value = "Férier"
    reference_ws["AN3"].fill = PatternFill(start_color = "000040", end_color = "000040", fill_type = "solid")
    reference_ws["AN3"].font = Font(name = "Arial", sz = 8, color = "FFFFFF")
    reference_ws["AN3"].alignment = Alignment("center", "center")

    reference_ws["AN4"].value = "Pond"
    reference_ws["AN4"].fill = PatternFill(start_color = "000040", end_color = "000040", fill_type = "solid")
    reference_ws["AN4"].font = Font(name = "Arial", sz = 8, color = "FFFFFF")
    reference_ws["AN4"].alignment = Alignment("center", "center")

    reference_ws["AN5"].value = "Congés entreprise"
    reference_ws["AN5"].fill = PatternFill(start_color = "000040", end_color = "000040", fill_type = "solid")
    reference_ws["AN5"].font = Font(name = "Arial", sz = 8, color = "FFFFFF")
    reference_ws["AN5"].alignment = Alignment("center", "center")

    Internal_formating(Worksheet = reference_ws, Range = "AN2:AR5", Type_border = "hair")
    Bottom_border_formating(Worksheet = reference_ws, Range = "AN2:AR2", Type_border = "dotted")
    Left_border_formating(Worksheet = reference_ws, Range = "AR2:AR5", Type_border = "dotted")
    Right_border_formating(Worksheet = reference_ws, Range = "AN2:AN5", Type_border = "dotted")
    External_formatting(Worksheet = reference_ws, Range = "AN2:AR5", Type_border = "thin")

    #Time calendar panel
    reference_ws["AN7"].value = "Année"
    reference_ws["AN7"].fill = PatternFill(start_color = "000040", end_color = "000040", fill_type = "solid")
    reference_ws["AN7"].font = Font(name = "Arial", sz = 8, color = "FFFFFF")
    reference_ws["AN7"].alignment = Alignment("center", "center")

    reference_ws["AN8"].value = "Mois"
    reference_ws["AN8"].fill = PatternFill(start_color = "000040", end_color = "000040", fill_type = "solid")
    reference_ws["AN8"].font = Font(name = "Arial", sz = 8, color = "FFFFFF")
    reference_ws["AN8"].alignment = Alignment("center", "center")
    
    reference_ws["AN9"].value = "Mois"
    reference_ws["AN9"].fill = PatternFill(start_color = "000040", end_color = "000040", fill_type = "solid")
    reference_ws["AN9"].font = Font(name = "Arial", sz = 8, color = "FFFFFF")
    reference_ws["AN9"].alignment = Alignment("center", "center")

    reference_ws["AO7"].value = "2023"
    reference_ws["AO8"].value = '=MONTH(TODAY())'
    reference_ws["AO9"].value ='=UPPER(TEXT(DATE(AO7,AO8,1),"mmmm"))'

    Internal_formating(Worksheet = reference_ws, Range = "AN7:AO9",Type_border = "hair")
    Right_border_formating(Worksheet = reference_ws, Range = "AN7:AN9", Type_border = "dotted")
    External_formatting(Worksheet = reference_ws, Range = "AN7:AO9", Type_border = "thin")

    #Pieces panel
    reference_ws["AN11"].value = "Types Piéces"
    reference_ws["AN11"].fill = PatternFill(start_color = "000040", end_color = "000040", fill_type = "solid")
    reference_ws["AN11"].font = Font(name = "Arial", sz = 8, color = "FFFFFF")
    reference_ws["AN11"].alignment = Alignment("center", "center")

    reference_ws["AN12"].value = "Catégories Piéces"
    reference_ws["AN12"].fill = PatternFill(start_color = "000040", end_color = "000040", fill_type = "solid")
    reference_ws["AN12"].font = Font(name = "Arial", sz = 8, color = "FFFFFF")
    reference_ws["AN12"].alignment = Alignment("center", "center")

    piece_type_list = ""
    category_piece_list = ""

    Range_splitted = 'AT3:AT8'.split(":")
    current_cell = Range_splitted[0]
    while current_cell != Range_splitted[1] : 
        if current_cell == Range_splitted[0] : 
            piece_type_list = "\'\"" + reference_ws[current_cell].value
            current_cell = vertical_range_travel(current_cell + ":" + Range_splitted[1])
        else :
            piece_type_list = piece_type_list + ", " + reference_ws[current_cell].value
            current_cell = vertical_range_travel(current_cell + ":" + Range_splitted[1])

    Range_splitted = 'AU3:AU12'.split(":")
    current_cell = Range_splitted[0]
    while current_cell != Range_splitted[1] : 
        if current_cell == Range_splitted[0] :
            category_piece_list =  "\'\"" + reference_ws[current_cell].value
            current_cell = vertical_range_travel(current_cell + ":" + Range_splitted[1])
        else :
            category_piece_list = category_piece_list + ", " + reference_ws[current_cell].value
            current_cell = vertical_range_travel(current_cell + ":" + Range_splitted[1])

    piece_type_list += "\"\'"
    category_piece_list += "\"\'"

    #print(piece_type_list)
    #print(category_piece_list)


    #paste piece_type_list to formula1 make bug Excel, found why
    piece_type_data_validation = DataValidation(type = 'list', formula1 = '"*, Facture, Avoir, Provisions, Contre passassion"', allow_blank = True)
    reference_ws.add_data_validation(piece_type_data_validation)
    piece_type_data_validation.add("AO11:AO11")

    #paste category_piece_list to formula1 make bug Excel, found why
    category_piece_data_validation = DataValidation(type = 'list', formula1 = '"*, Interim, Pole Sante, Eau, Gaz, Electricite, Cantine, Nettoyage, Transport"', allow_blank = True)
    reference_ws.add_data_validation(category_piece_data_validation)
    category_piece_data_validation.add("AO12:AO12")

    reference_ws.merge_cells("BA1:BB1")
    reference_ws["BA1"].value = "Sites"
    reference_ws["BA1"].fill = PatternFill(start_color = "000040", end_color = "000040", fill_type = "solid")
    reference_ws["BA1"].font = Font(name = "Arial", sz = 8, color = "FFFFFF")
    reference_ws["BA1"].alignment = Alignment("center", "center")

    reference_ws["BA2"].value = "Centre de cout"
    reference_ws["BA2"].fill = PatternFill(start_color = "000040", end_color = "000040", fill_type = "solid")
    reference_ws["BA2"].font = Font(name = "Arial", sz = 8, color = "FFFFFF")
    reference_ws["BA2"].alignment = Alignment("center", "center")

    reference_ws["BB2"].value = "Libelle"
    reference_ws["BB2"].fill = PatternFill(start_color = "000040", end_color = "000040", fill_type = "solid")
    reference_ws["BB2"].font = Font(name = "Arial", sz = 8, color = "FFFFFF")
    reference_ws["BB2"].alignment = Alignment("center", "center")

    Bottom_border_formating(Worksheet = reference_ws, Range = "BA1:BB1", Type_border = "dotted")  
    Right_border_formating(Worksheet = reference_ws, Range = "BA2:BA2", Type_border = "hair")  
    External_formatting(Worksheet = reference_ws, Range = "BA1:BB2", Type_border = "thin")

    
    reference_ws["BA3"].value = "6200"
    reference_ws["BA4"].value = "6201"
    reference_ws["BA5"].value = "6202"
    reference_ws["BA6"].value = "6203"
    reference_ws["BA7"].value = "6204"
    reference_ws["BA8"].value = "6205"
    reference_ws["BA9"].value = "6206"
    reference_ws["BA10"].value = "6207"
    reference_ws["BA11"].value = "6208"
    
    reference_ws["BB3"].value = "Usine"
    reference_ws["BB4"].value = "Montage"
    reference_ws["BB5"].value = "Usinage"
    reference_ws["BB6"].value = "Magasin"
    reference_ws["BB7"].value = "Expedition"
    reference_ws["BB8"].value = "Reception"
    reference_ws["BB9"].value = "Restauration"
    reference_ws["BB10"].value = "Propreté"
    reference_ws["BB11"].value = "Logistique"


    reference_ws.merge_cells("BF1:BH1")
    reference_ws["BF1"].value = "Comptes comptable"
    reference_ws["BF1"].fill = PatternFill(start_color = "000040", end_color = "000040", fill_type = "solid")
    reference_ws["BF1"].font = Font(name = "Arial", sz = 8, color = "FFFFFF")
    reference_ws["BF1"].alignment = Alignment("center", "center")
    
    reference_ws["BF2"].value = "Comptes"
    reference_ws["BF2"].fill = PatternFill(start_color = "000040", end_color = "000040", fill_type = "solid")
    reference_ws["BF2"].font = Font(name = "Arial", sz = 8, color = "FFFFFF")
    reference_ws["BF2"].alignment = Alignment("center", "center")
    
    reference_ws["BG2"].value = "Libelle"
    reference_ws["BG2"].fill = PatternFill(start_color = "000040", end_color = "000040", fill_type = "solid")
    reference_ws["BG2"].font = Font(name = "Arial", sz = 8, color = "FFFFFF")
    reference_ws["BG2"].alignment = Alignment("center", "center")
    
    reference_ws["BH2"].value = "Abreviation"
    reference_ws["BH2"].fill = PatternFill(start_color = "000040", end_color = "000040", fill_type = "solid")
    reference_ws["BH2"].font = Font(name = "Arial", sz = 8, color = "FFFFFF")
    reference_ws["BH2"].alignment = Alignment("center", "center")

    Internal_formating(Worksheet = reference_ws, Range = "BF1:BH2", Type_border = "hair")
    External_formatting(Worksheet = reference_ws,  Range = "BF1:BH2", Type_border = "thin")
    Bottom_border_formating(Worksheet = reference_ws,  Range = "BF1:BH1", Type_border = "dotted")

    reference_ws["BF3"].value = "62110000"
    reference_ws["BF4"].value = "64750000"
    reference_ws["BF5"].value = "60610001"
    reference_ws["BF6"].value = "60610002"
    reference_ws["BF7"].value = "60610003"
    reference_ws["BF8"].value = "61100000"
    reference_ws["BF9"].value = "61520000"
    reference_ws["BF10"].value = "62420000"

    reference_ws["BG3"].value = "Personnel intérimaire"
    reference_ws["BG4"].value = "Médecine du travail, pharmacie "
    reference_ws["BG5"].value = "Fournitures non stockables (eau, énergie) - Eau"
    reference_ws["BG6"].value = "Fournitures non stockables (eau, énergie) - Gaz"
    reference_ws["BG7"].value = "Fournitures non stockables (eau, énergie) - Electricité"
    reference_ws["BG8"].value = "Sous-traitance générale - Cantine"
    reference_ws["BG9"].value = "Entretien et réparations sur biens immobiliers"
    reference_ws["BG10"].value = "Transports sur ventes"

    reference_ws["BH3"].value = "Interim"
    reference_ws["BH4"].value = "Pole sante"
    reference_ws["BH5"].value = "Eau"
    reference_ws["BH6"].value = "Gaz"
    reference_ws["BH7"].value = "Electricite"
    reference_ws["BH8"].value = "Cantine"
    reference_ws["BH9"].value = "Nettoyage"
    reference_ws["BH10"].value = "Transport"


    reference_ws.merge_cells("BL1:BM1")
    reference_ws["BL1"].value = "Fournisseur"
    reference_ws["BL1"].fill = PatternFill(start_color = "000040", end_color = "000040", fill_type = "solid")
    reference_ws["BL1"].font = Font(name = "Arial", sz = 8, color = "FFFFFF")
    reference_ws["BL1"].alignment = Alignment("center", "center")
    
    reference_ws["BL2"].value = "N° Fournisseur"
    reference_ws["BL2"].fill = PatternFill(start_color = "000040", end_color = "000040", fill_type = "solid")
    reference_ws["BL2"].font = Font(name = "Arial", sz = 8, color = "FFFFFF")
    reference_ws["BL2"].alignment = Alignment("center", "center")
    
    reference_ws["BM2"].value = "Designation Fournisseur"
    reference_ws["BM2"].fill = PatternFill(start_color = "000040", end_color = "000040", fill_type = "solid")
    reference_ws["BM2"].font = Font(name = "Arial", sz = 8, color = "FFFFFF")
    reference_ws["BM2"].alignment = Alignment("center", "center")

    Internal_formating(Worksheet = reference_ws, Range = "BL1:BM2", Type_border = "hair")
    External_formatting(Worksheet = reference_ws, Range = "BL1:BM2", Type_border = "thin")
    Bottom_border_formating(Worksheet = reference_ws, Range = "BL1:BM1", Type_border = "dotted")

    reference_ws["BL3"].value = "20001"
    reference_ws["BL4"].value = "20002"
    reference_ws["BL5"].value = "20003"
    reference_ws["BL6"].value = "20004"
    reference_ws["BL7"].value = "20005"
    reference_ws["BL8"].value = "20006"
    reference_ws["BL9"].value = "20007"
    reference_ws["BL10"].value = "20008"

    reference_ws["BM3"].value = "Agence interim 1"
    reference_ws["BM4"].value = "Agence interim 2"
    reference_ws["BM5"].value = "Medecine du travail"
    reference_ws["BM6"].value = "Eau de ville"
    reference_ws["BM7"].value = "EDF"
    reference_ws["BM8"].value = "GDF"
    reference_ws["BM9"].value = "Restaurateur"
    reference_ws["BM10"].value = "Nettoyage"


    Provisions_file.save(PROVISIONS_FILE_PATH)

Reference_worksheet()
